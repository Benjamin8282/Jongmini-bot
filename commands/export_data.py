import asyncio
import io
import zipfile
from collections import defaultdict
from datetime import datetime, timedelta, timezone

import discord
from discord import app_commands
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from core.db import get_all_watch_items, get_conn
from core.logger import logger

KST = timezone(timedelta(hours=9))

# 디스코드 일반 서버(부스트 1) 첨부 한도 25 MiB. 안전 마진 두고 24 MiB로 청크 크기 설정
CHUNK_SIZE = 24 * 1024 * 1024
MAX_FILES_PER_MESSAGE = 10

DATA_HEADERS = ["거래일시", "단가", "총액", "수량"]

# Excel 시트당 최대 행 수 1,048,576 (헤더 1행 차지)
MAX_ROWS_PER_SHEET = 1_048_575

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
LINK_FONT = Font(color="0563C1", underline="single", size=11)
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
NUM_FMT = '#,##0'


async def _fetch_export_data(item_id: str | None, days: int | None) -> tuple[list[dict], int]:
    """내보낼 데이터 조회. (rows, total_count) 반환."""
    conditions = []
    params = []

    if item_id:
        conditions.append("aph.item_id = ?")
        params.append(item_id)
    if days:
        cutoff = (datetime.now(KST) - timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
        conditions.append("aph.sold_date >= ?")
        params.append(cutoff)

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    conn = await get_conn()
    cursor = await conn.execute(
        f"SELECT COUNT(*) as cnt FROM auction_price_history aph {where}", params
    )
    total = (await cursor.fetchone())["cnt"]

    cursor = await conn.execute(f"""
        SELECT aph.item_id, aw.item_name, aph.sold_date,
               aph.unit_price, aph.price, aph.count
        FROM auction_price_history aph
        LEFT JOIN auction_watch_items aw ON aph.item_id = aw.item_id
        {where}
        ORDER BY aw.item_name, aph.sold_date
    """, params)
    rows = await cursor.fetchall()

    return [dict(r) for r in rows], total


def _safe_sheet_name(name: str) -> str:
    """시트 이름에 사용할 수 없는 문자 제거 (최대 31자)."""
    for ch in r'\/:*?[]':
        name = name.replace(ch, "_")
    return name[:31]


def _style_header(ws, col_count: int):
    """헤더 행 스타일 적용."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def _auto_width(ws, col_count: int, data_row_count: int):
    """컬럼 너비 자동 조정."""
    for col in range(1, col_count + 1):
        max_len = len(str(ws.cell(row=1, column=col).value or ""))
        for row in range(2, min(data_row_count + 2, 102)):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 30)


def _write_data_sheet(ws, rows: list[dict]):
    """데이터 시트에 행 쓰기."""
    ws.append(DATA_HEADERS)
    _style_header(ws, len(DATA_HEADERS))

    for row in rows:
        ws.append([
            row["sold_date"],
            row["unit_price"],
            row["price"],
            row["count"],
        ])

    for r in range(2, len(rows) + 2):
        for c in [2, 3, 4]:
            ws.cell(row=r, column=c).number_format = NUM_FMT
        ws.cell(row=r, column=1).border = THIN_BORDER

    _auto_width(ws, len(DATA_HEADERS), len(rows))
    ws.freeze_panes = "A2"


def _write_summary_sheet(ws, item_stats: list[dict], export_time: str, days: int | None):
    """요약 시트 작성 (통계 + 아이템별 시트 하이퍼링크)."""
    ws.append(["경매장 시세 데이터 내보내기"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws.merge_cells("A1:E1")

    ws.append([])
    ws.append(["내보내기 시간", export_time])
    ws.cell(row=3, column=1).font = Font(bold=True)

    period = f"최근 {days}일" if days else "전체 기간"
    ws.append(["기간", period])
    ws.cell(row=4, column=1).font = Font(bold=True)

    total_records = sum(s["count"] for s in item_stats)
    ws.append(["총 거래 건수", total_records])
    ws.cell(row=5, column=1).font = Font(bold=True)
    ws.cell(row=5, column=2).number_format = NUM_FMT

    ws.append(["아이템 종류", len(item_stats)])
    ws.cell(row=6, column=1).font = Font(bold=True)

    ws.append([])
    ws.append(["아이템", "거래 건수", "평균 단가", "최저 단가", "최고 단가"])
    _style_header(ws, 5)
    header_row = ws.max_row

    for i, s in enumerate(item_stats):
        row_num = header_row + 1 + i
        sheet_name = s["sheet_name"]

        cell = ws.cell(row=row_num, column=1, value=s["name"])
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = LINK_FONT

        ws.cell(row=row_num, column=2, value=s["count"]).number_format = NUM_FMT
        ws.cell(row=row_num, column=3, value=s["avg_price"]).number_format = NUM_FMT
        ws.cell(row=row_num, column=4, value=s["min_price"]).number_format = NUM_FMT
        ws.cell(row=row_num, column=5, value=s["max_price"]).number_format = NUM_FMT

    _auto_width(ws, 5, len(item_stats) + header_row)
    ws.sheet_properties.tabColor = "4472C4"


def _build_xlsx(rows: list[dict], export_time: str, days: int | None) -> bytes:
    """아이템별 시트 + 요약 시트가 포함된 Excel 파일 생성."""
    grouped = defaultdict(list)
    for row in rows:
        name = row.get("item_name") or row.get("item_id", "unknown")
        row["item_name"] = name
        grouped[name].append(row)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "요약"

    item_stats = []
    for name, item_rows in sorted(grouped.items()):
        if len(item_rows) <= MAX_ROWS_PER_SHEET:
            sheet_name = _safe_sheet_name(name)
            ws = wb.create_sheet(title=sheet_name)
            _write_data_sheet(ws, item_rows)
            first_sheet_name = sheet_name
        else:
            # 시트당 행 한도 초과 시 분할
            n_chunks = (len(item_rows) + MAX_ROWS_PER_SHEET - 1) // MAX_ROWS_PER_SHEET
            first_sheet_name = None
            for i in range(n_chunks):
                chunk = item_rows[i * MAX_ROWS_PER_SHEET:(i + 1) * MAX_ROWS_PER_SHEET]
                suffix = f"_{i + 1}"
                base = _safe_sheet_name(name)[:31 - len(suffix)]
                chunk_sheet_name = f"{base}{suffix}"
                ws = wb.create_sheet(title=chunk_sheet_name)
                _write_data_sheet(ws, chunk)
                if first_sheet_name is None:
                    first_sheet_name = chunk_sheet_name

        prices = [r["unit_price"] for r in item_rows if r["unit_price"] > 0]
        item_stats.append({
            "name": name,
            "sheet_name": first_sheet_name,
            "count": len(item_rows),
            "avg_price": round(sum(prices) / len(prices)) if prices else 0,
            "min_price": min(prices) if prices else 0,
            "max_price": max(prices) if prices else 0,
        })

    _write_summary_sheet(ws_summary, item_stats, export_time, days)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _zip_xlsx(xlsx_bytes: bytes, inner_filename: str) -> bytes:
    """xlsx 파일을 zip으로 압축."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        zf.writestr(inner_filename, xlsx_bytes)
    return buf.getvalue()


def _split_into_chunks(data: bytes, chunk_size: int) -> list[bytes]:
    """바이트를 chunk_size 크기로 분할."""
    return [data[i:i + chunk_size] for i in range(0, len(data), chunk_size)]


def _format_size(size_bytes: int) -> str:
    """바이트를 읽기 좋은 크기로 변환."""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    if size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes / (1024 * 1024):.1f} MB"


@app_commands.command(
    name="데이터내보내기",
    description="경매장 시세 데이터를 분할 압축해 디스코드에 업로드합니다"
)
@app_commands.describe(
    days="최근 N일 데이터만 (미입력 시 전체)",
    item="특정 아이템만 (미입력 시 전체)"
)
async def export_data(interaction: discord.Interaction,
                      days: int | None = None,
                      item: str | None = None):
    await interaction.response.defer(thinking=True)

    try:
        item_id = None
        item_display = "전체"
        if item:
            watch_items = await get_all_watch_items()
            matched = [w for w in watch_items if item in w["item_name"]]
            if not matched:
                await interaction.followup.send(f"'{item}'에 해당하는 감시 아이템을 찾을 수 없습니다.")
                return
            if len(matched) > 1:
                names = "\n".join(f"- {w['item_name']}" for w in matched)
                await interaction.followup.send(
                    f"여러 아이템이 매칭됩니다. 정확한 이름을 입력해주세요:\n{names}"
                )
                return
            item_id = matched[0]["item_id"]
            item_display = matched[0]["item_name"]

        rows, total = await _fetch_export_data(item_id, days)

        if not rows:
            await interaction.followup.send("내보낼 데이터가 없습니다.")
            return

        now = datetime.now(KST)
        timestamp = now.strftime("%Y%m%d_%H%M%S")
        export_time = now.strftime("%Y-%m-%d %H:%M:%S KST")

        xlsx_filename = f"auction_data_{timestamp}.xlsx"
        zip_filename = f"auction_data_{timestamp}.zip"

        xlsx_bytes = await asyncio.to_thread(_build_xlsx, rows, export_time, days)
        zip_bytes = await asyncio.to_thread(_zip_xlsx, xlsx_bytes, xlsx_filename)
        chunks = _split_into_chunks(zip_bytes, CHUNK_SIZE)

        size = _format_size(len(zip_bytes))
        period = f"최근 {days}일" if days else "전체 기간"
        item_count = len(set(
            (r.get("item_name") or r.get("item_id")) for r in rows
        ))

        embed = discord.Embed(
            title="데이터 내보내기 완료",
            description=(
                f"총 **{len(chunks)}개 파트**로 분할되었습니다. "
                f"모든 파트를 받은 뒤 아래 명령으로 합쳐주세요.\n\n"
                f"**Windows (cmd)**\n"
                f"```\ncopy /b {zip_filename}.* {zip_filename}\n```\n"
                f"**Linux / macOS**\n"
                f"```\ncat {zip_filename}.* > {zip_filename}\n```\n"
                f"합친 후 zip을 풀면 `{xlsx_filename}` 가 나옵니다."
            ),
            color=discord.Color.green()
        )
        embed.add_field(name="아이템", value=item_display, inline=True)
        embed.add_field(name="기간", value=period, inline=True)
        embed.add_field(name="거래 건수", value=f"{total:,}건", inline=True)
        embed.add_field(name="zip 크기", value=size, inline=True)
        embed.add_field(name="시트", value=f"요약 + {item_count}개 아이템", inline=True)
        embed.add_field(name="분할 파트", value=f"{len(chunks)}개", inline=True)
        embed.set_footer(text=export_time)

        await interaction.followup.send(embed=embed)

        total_batches = (len(chunks) + MAX_FILES_PER_MESSAGE - 1) // MAX_FILES_PER_MESSAGE
        for batch_idx in range(total_batches):
            start = batch_idx * MAX_FILES_PER_MESSAGE
            batch = chunks[start:start + MAX_FILES_PER_MESSAGE]
            files = [
                discord.File(
                    io.BytesIO(chunk),
                    filename=f"{zip_filename}.{start + i + 1:03d}"
                )
                for i, chunk in enumerate(batch)
            ]
            content = (
                f"파트 {start + 1}~{start + len(batch)} / {len(chunks)} "
                f"(메시지 {batch_idx + 1}/{total_batches})"
            )
            await interaction.followup.send(content=content, files=files)

        logger.info(
            f"데이터 내보내기 완료: {zip_filename} "
            f"({total}건, {size}, {len(chunks)}개 파트)"
        )

    except Exception as e:
        await interaction.followup.send(
            "내보내기 중 오류가 발생했습니다. 로그를 확인해주세요.", ephemeral=True
        )
        logger.error(f"데이터 내보내기 실패: {e}", exc_info=True)
